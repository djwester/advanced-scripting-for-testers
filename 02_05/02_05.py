from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

manual_testing_wb = load_workbook("manual_testing_results.xlsx", data_only=True)

manual_testing_sheet = manual_testing_wb.active


def find_column_index(header_row, header_name):
    for header in header_row:
        if header.value == header_name:
            return header.column


def get_failure_rates(worksheet, date_col_index, results_col_index, daily_failures):
    date_mapping = defaultdict(list)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        day = datetime.date(row[date_col_index - 1])
        date_mapping[day].append(row[results_col_index - 1])

    for day, results in date_mapping.items():
        failure_rate = results.count("FAIL") / len(results) if results else 0
        daily_failures[day].append(failure_rate)
    return daily_failures


header_row = manual_testing_sheet[1]
results_column_index = find_column_index(header_row, "Result")
date_col_index = find_column_index(header_row, "Date")

daily_failures = defaultdict(list)
daily_failures = get_failure_rates(
    manual_testing_sheet, date_col_index, results_column_index, daily_failures
)

test_automation_wb = load_workbook("test_automation_results.xlsx", data_only=True)
test_automation_sheet = test_automation_wb.active

header_row = test_automation_sheet[1]
status_column_index = find_column_index(header_row, "Status")
timestamp_col_index = find_column_index(header_row, "Timestamp")

daily_failures = get_failure_rates(
    test_automation_sheet,
    timestamp_col_index,
    status_column_index,
    daily_failures,
)

wb = Workbook()

final_report_worksheet = wb.active
headers = ["Date", "Manual Result", "Automation Result"]

final_report_worksheet.append(headers)
for day, failures in daily_failures.items():
    final_report_worksheet.append([day, failures[0], failures[1]])

chart = LineChart()
chart.title = "Daily Failure Rates"
chart.x_axis.title = "Date"
chart.y_axis.title = "Failure Rate"
reference = Reference(
    final_report_worksheet,
    min_col=2,
    min_row=1,
    max_col=3,
    max_row=final_report_worksheet.max_row,
)
chart.add_data(reference, titles_from_data=True)
final_report_worksheet.add_chart(chart, "E2")

file_path = "final_report.xlsx"
wb.save(file_path)
