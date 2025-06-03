from openpyxl import load_workbook

workbook = load_workbook("dummy_data.xlsx", data_only=True)

sheet = workbook.active

for row in sheet.iter_rows():
    for cell in row:
        print(cell.value)
