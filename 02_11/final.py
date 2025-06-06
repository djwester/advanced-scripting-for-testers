from collections import defaultdict

import gspread
import openpyxl
from google.oauth2.service_account import Credentials

SERVICE_ACCOUNT_FILE = "../advancedscripting-da4417eb3e72.json"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]
creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
client = gspread.authorize(creds)

g_sheet = client.open_by_url(
    "https://docs.google.com/spreadsheets/d/1S_1difkV4ZirGJ7jO_yCcvVqW0_YyY1SS30qoWp8Iek/edit?gid=0#gid=0"
)
worksheet = g_sheet.get_worksheet(0)

date_mapping = defaultdict(int)
for row in worksheet.col_values(5)[1:]:
    date_mapping[row] += 1

workbook = openpyxl.load_workbook("final_report.xlsx")
sheet = workbook.active

sheet.cell(1, 4, value="Bug Count")
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
    date_value = row[0]
    bug_count = date_mapping[date_value]
    sheet.cell(i + 2, 4, value=bug_count)

workbook.save("final_report.xlsx")
