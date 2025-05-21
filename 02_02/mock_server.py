import sqlite3

from fastapi import FastAPI, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel

app = FastAPI()

connection = sqlite3.connect(":memory:")


# Define the request payload structure
class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/api/login")
async def login(request: LoginRequest):
    # Mock response logic
    if not request.password == "test123":
        raise HTTPException(status_code=401, detail="Invalid credentials")
    cursor = connection.cursor()
    cursor.execute(
        "SELECT is_active FROM users WHERE username = ?",
        (request.username,),
    )
    result = cursor.fetchone()
    if result[0]:
        return {"message": f"Welcome, {request.username}!"}
    else:
        raise HTTPException(status_code=401, detail="User is inactive")


def create_in_memory_db():
    cursor = connection.cursor()
    cursor.execute("""
        CREATE TABLE users (
            username TEXT PRIMARY KEY,
            password TEXT,
            is_active INTEGER,
            login_attempts INTEGER DEFAULT 0,
            last_login TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            login_count INTEGER DEFAULT 0
        )
    """)
    return connection


def read_sample_file_into_memory():
    workbook = load_workbook(filename="sample_test_data.xlsx", data_only=True)

    sheet = workbook.active

    # populate the in-memory database with sample data
    db_connection = create_in_memory_db()
    cursor = db_connection.cursor()
    for row in sheet.iter_rows(values_only=True):
        cursor.execute(
            "INSERT INTO users (username, password, is_active, login_attempts, last_login, login_count) VALUES (?, ?, ?, ?, ?, ?)",
            (row[1], "test123", row[5], row[7], row[8], row[6]),
        )


if __name__ == "__main__":
    read_sample_file_into_memory()
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8002)
